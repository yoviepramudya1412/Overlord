from os import truncate
from django.db.models.functions import TruncDate
import time
from django.conf import settings
from django.shortcuts import redirect, render
from requests import request
from .forms import *
from .models import *
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect
from django.http import JsonResponse
from django.core.serializers import serialize
from django.shortcuts import get_object_or_404
from django.db.models import Count
from datetime import datetime, timedelta
from django.db.models.functions import TruncDay
from django.db.models.functions import TruncMonth
from collections import defaultdict
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter





# fitur tambahan

def export_kerusakan_to_excel(request):
    kerusakan_data = Kerusakan.objects.all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="kerusakan_data_FPJ.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Kerusakan Data'

    headers = [ 'Tanggal Laporkan', 'Nama Masyarakat','No Telepon', 'Nama Fasilitas', 'Jenis Perlengkapan', 'Rusak']

    # Set title
    title_cell = ws.cell(row=1, column=1, value="Data Kerusakan")
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.font = Font(size=16, bold=True)

    # Set headers
    ws.append(headers)

    max_lengths = [len(header) for header in headers]  # Initialize max_lengths list

    for kerusakan in kerusakan_data:
        row = [
            
            kerusakan.tanggal_laporkan.strftime('%Y-%m-%d') if kerusakan.tanggal_laporkan else '',
            kerusakan.masyarakatid.nama,
            kerusakan.masyarakatid.notelepon,
            kerusakan.nama_fasilitas.nama_fasilitas if kerusakan.nama_fasilitas else '',
            kerusakan.jenis_perlengkapan.jenis_perlengkapan if kerusakan.jenis_perlengkapan else '',
            kerusakan.rusak.tiperusak ,
        ]
        ws.append(row)

        # Update max_lengths based on current row data
        for i, cell_value in enumerate(row):
            max_lengths[i] = max(max_lengths[i], len(str(cell_value)))

    # Set column widths based on max_lengths
    for i, max_length in enumerate(max_lengths):
        ws.column_dimensions[chr(65 + i)].width = max_length + 2  # +2 for some padding

    # Set borders for all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    wb.save(response)
    return response




def export_pengajuan_to_excel(request):
    pengajuan_data = Pengajuan.objects.all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pengajuan_data_FPJ.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pengajuan Data'

    headers = [ 'Tanggal Ajukan', 'Nama Masyarakat', 'Nomor Telepon','Nama Fasilitas','Jenis Perlengkapan','Fasilitas Khusus','Latitude','Longitude','Status']

    # Set title
    title_cell = ws.cell(row=1, column=1, value="Data Pengajuan")
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.font = Font(size=16, bold=True)

    # Set headers
    ws.append(headers)

    max_lengths = [len(header) for header in headers]  # Initialize max_lengths list

    for pengajuan in pengajuan_data:
        row = [
            
            pengajuan.tanggal_ajukan.strftime('%Y-%m-%d') if pengajuan.tanggal_ajukan else '',
            pengajuan.masyarakatid.nama,
            pengajuan.masyarakatid.notelepon,
            pengajuan.nama_fasilitas.nama_fasilitas if pengajuan.nama_fasilitas else '',
            pengajuan.jenis_perlengkapan.jenis_perlengkapan if pengajuan.jenis_perlengkapan else '',
            pengajuan.Fasilitas_khusus,
            pengajuan.location.latitude if pengajuan.location else '',
            pengajuan.location.longitude if pengajuan.location else '',
            pengajuan.statuspengajuan.tipestatus if pengajuan.statuspengajuan else '',
        ]
        ws.append(row)

        # Update max_lengths based on current row data
        for i, cell_value in enumerate(row):
            max_lengths[i] = max(max_lengths[i], len(str(cell_value)))

    # Set column widths based on max_lengths
    for i, max_length in enumerate(max_lengths):
        ws.column_dimensions[chr(65 + i)].width = max_length + 2  # +2 for some padding

    # Set borders for all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    wb.save(response)
    return response




def export_to_excel(request):
    pembangunan_data = Pembangunan.objects.filter(status__tipestatus='PEMBANGUNAN')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pembangunan_data_FPJ.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pembangunan Data'

    headers = ['Tanggal Bangun', 'Konstruksi Selesai', 'Ruas Jalan', 'Nama Fasilitas', 'Jenis Perlengkapan', 'Kondisi', 'Volume', 'Latitude', 'Longitude', 'Status']

    # Set title
    title_cell = ws.cell(row=1, column=1, value="Pembangunan Alat Fasilitas Perlengkapan Aceh")
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.font = Font(size=16, bold=True)

    # Set headers
    ws.append(headers)

    for pembangunan in pembangunan_data:
        row = [
            pembangunan.tanggal_bangun.strftime('%Y-%m-%d') if pembangunan.tanggal_bangun else '',
            pembangunan.konstruksi_selesai.strftime('%Y-%m-%d') if pembangunan.konstruksi_selesai else '',
            pembangunan.ruasjalan,
            pembangunan.nama_fasilitas.nama_fasilitas if pembangunan.nama_fasilitas else '',
            pembangunan.jenis_perlengkapan.jenis_perlengkapan if pembangunan.jenis_perlengkapan else '',
            pembangunan.kondisi.tipekondisi if pembangunan.kondisi else '',
            pembangunan.volume if pembangunan.volume else '',
            pembangunan.location.latitude if pembangunan.location else '',
            pembangunan.location.longitude if pembangunan.location else '',
            pembangunan.status.tipestatus if pembangunan.status else '',
        ]
        ws.append(row)

    # Set borders for all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    wb.save(response)
    return response











def export_to_excelper(request):
    pembangunan_data = Pembangunan.objects.filter(status__tipestatus='PERENCANAAN')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="perencanaan_data_FPJ.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pembangunan Data'

    headers = ['Tanggal Bangun', 'Konstruksi Selesai', 'Ruas Jalan', 'Nama Fasilitas', 'Jenis Perlengkapan', 'Kondisi', 'Volume', 'Latitude', 'Longitude', 'Status']

    # Set title
    title_cell = ws.cell(row=1, column=1, value="Pembangunan Alat Fasilitas Perlengkapan Aceh")
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.font = Font(size=16, bold=True)

    # Set headers
    ws.append(headers)

    for pembangunan in pembangunan_data:
        row = [
            pembangunan.tanggal_bangun.strftime('%Y-%m-%d') if pembangunan.tanggal_bangun else '',
            pembangunan.konstruksi_selesai.strftime('%Y-%m-%d') if pembangunan.konstruksi_selesai else '',
            pembangunan.ruasjalan,
            pembangunan.nama_fasilitas.nama_fasilitas if pembangunan.nama_fasilitas else '',
            pembangunan.jenis_perlengkapan.jenis_perlengkapan if pembangunan.jenis_perlengkapan else '',
            pembangunan.kondisi.tipekondisi if pembangunan.kondisi else '',
            pembangunan.volume if pembangunan.volume else '',
            pembangunan.location.latitude if pembangunan.location else '',
            pembangunan.location.longitude if pembangunan.location else '',
            pembangunan.status.tipestatus if pembangunan.status else '',
        ]
        ws.append(row)

    # Set borders for all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    wb.save(response)
    return response



# Tampilan statistik
@login_required(login_url=settings.LOGIN_URL)
def statistikfasiltas(request):
    # data = Fasilitas_perlengkapan.objects.annotate(date_count=Count('tanggal_ditambahkan')).values('tanggal_ditambahkan', 'date_count')
    # chart_data = [{'x': entry['tanggal_ditambahkan'].strftime('%Y-%m-%d'), 'y': entry['date_count']} for entry in data]

    data = Fasilitas_perlengkapan.objects.annotate(date_count=Count('tanggal_ditambahkan')).values('tanggal_ditambahkan', 'date_count')
    
    sorted_data = sorted(data, key=lambda entry: entry['tanggal_ditambahkan'])

    data_bar = {}
    for entry in sorted_data:
        date = entry['tanggal_ditambahkan'].strftime('%Y-%m-%d')
        if date in data_bar:
            data_bar[date] += entry['date_count']
        else:
            data_bar[date] = entry['date_count']
        
    chart_data = [{'x': datetime.strptime(date, '%Y-%m-%d').strftime('%d %b %Y'), 'y': count} for date, count in data_bar.items()]

    


    total_entries = Fasilitas_perlengkapan.objects.count()
    max_entries = 50  # Maksimal fasilitas per tahun

    # Hitung persentase pertambahan data
    percentage_increase = (total_entries / max_entries) * 100
    
    
    radialGradientOptions = {
        
        'series': [percentage_increase],
        'chart': {
            'height': 350,
            'type': 'radialBar',
            'toolbar': {
                'show': 'true',
            },
        },
        'plotOptions': {
        'radialBar': {
            'startAngle': -135,
            'endAngle': 225,
            'hollow': {
                'margin': 0,
                'size': '70%',
                'background': '#fff',
                
                'imageOffsetX': 0,
                'imageOffsetY': 0,
                'position': 'front',
                'dropShadow': {
                    'enabled': 'true',
                    'top': 3,
                    'left': 0,
                    'blur': 4,
                    'opacity': 0.24,
                },
            },
            'track': {
                'background': '#fff',
                'strokeWidth': '67%',
                'margin': 0,  # margin is in pixels
                'dropShadow': {
                    'enabled': 'true',
                    'top': -3,
                    'left': 0,
                    'blur': 4,
                    'opacity': 0.35,
                },
            },
            
        },
    },
        
        
        'fill': {
            'type': 'gradient',
            'gradient': {
                'shade': 'dark',
                'type': 'horizontal',
                'shadeIntensity': 0.5,
                'gradientToColors': ['#ABE5A1'],
                'inverseColors': 'true',
                'opacityFrom': 1,
                'opacityTo': 1,
                'stops': [0, 100],
            },
        },
        'stroke': {
            'lineCap': 'round',
        },
        'labels': ['Percent'],
    }

    data_area = Fasilitas_perlengkapan.objects.values('jenis_perlengkapan', 'tanggal_ditambahkan').annotate(data_count=Count('jenis_perlengkapan'))

    data_linearea = defaultdict(dict)
    for entry in data_area:
        jenis_perlengkapan = entry['jenis_perlengkapan']
        tanggal_ditambahkan = entry['tanggal_ditambahkan'].strftime('%Y-%m-%d')
        data_count = entry['data_count']

        data_linearea[jenis_perlengkapan][tanggal_ditambahkan] = data_count

    categories = sorted(set([tanggal for data in data_linearea.values() for tanggal in data.keys()]))

    series = [{'name': jenis, 'data': [data.get(tanggal, 0) for tanggal in categories]} for jenis, data in data_linearea.items()]

    chart_area = {
        'categories': categories,
        'series': series,
    }
    

    context = { 
                'chart_area':chart_area,
                'data_linearea':data_linearea,
                'chart_data': chart_data,
                'radialGradientOptions': radialGradientOptions,
                }
    
    return render(request, 'statistik/statistik fasilitas.html',context)

@login_required(login_url=settings.LOGIN_URL)
def statistikkerusakan(request):
    rusak_choices = dict(Rusak.RUSAK_CHOICES)  # Mengakses RUSAK_CHOICES dari model Rusak

    data = Kerusakan.objects.values('rusak__tiperusak', 'tanggal_laporkan').annotate(data_count=Count('rusak__tiperusak'))

    data_linearea = defaultdict(lambda: defaultdict(int))  # Menggunakan defaultdict untuk menghitung total tiap tipe rusak
    for entry in data:
        rusak_tipe = entry['rusak__tiperusak']
        tanggal_laporkan = entry['tanggal_laporkan'].strftime('%Y-%m-%d')
        data_count = entry['data_count']

        data_linearea[rusak_tipe][tanggal_laporkan] += data_count

    chart_categories = sorted(set([tanggal for data in data_linearea.values() for tanggal in data.keys()]))

    chart_data = [{'name': rusak_choices[rusak], 'data': [data.get(tanggal, 0) for tanggal in chart_categories]} for rusak, data in data_linearea.items()]


    # untuk chart lain 
    total_entries = Kerusakan.objects.count()
    max_entries = 100  # Maksimal kerusakan per tahun

    # Hitung persentase pertambahan data
    percentage_increase = (total_entries / max_entries) * 100
    
    # untuk chart lain 
    data_inline = Kerusakan.objects.values('jenis_perlengkapan__jenis_perlengkapan', 'tanggal_laporkan').annotate(data_inline_count=Count('jenis_perlengkapan__jenis_perlengkapan'))

    data_line = defaultdict(lambda: defaultdict(int))
    for entry in data_inline:
        jenis_perlengkapan = entry['jenis_perlengkapan__jenis_perlengkapan']
        tanggal_laporkan = entry['tanggal_laporkan'].strftime('%Y-%m-%d')
        data_inline_count = entry['data_inline_count']

        data_line[jenis_perlengkapan][tanggal_laporkan] += data_inline_count

    categories = [{'name': jenis_perlengkapan, 'data': [data.get(tanggal, 0) for tanggal in chart_categories]} for jenis_perlengkapan, data in data_line.items()]

    chart_data_line = {
        'series1': categories,
        'categories1': chart_categories,
    }
    
    # charlain
    data_line = Kerusakan.objects.annotate(date_count=Count('tanggal_laporkan')).values('tanggal_laporkan', 'date_count').order_by('tanggal_laporkan')
    data_line_plot = {}
    for entry in data_line:
        date = entry['tanggal_laporkan'].strftime('%Y-%m-%d')
        if date in data_line_plot:
            data_line_plot[date] += entry['date_count']
        else:
            data_line_plot[date] = entry['date_count']
    chart_data_bar = [{'x': datetime.strptime(date, '%Y-%m-%d').strftime('%d %b %Y'), 'y': count} for date, count in data_line_plot.items()]
    
    context = {
        'chart_data_bar' : chart_data_bar,
        'chart_data_line': chart_data_line,
        'chart_data': {
            'series': chart_data,
            'categories': chart_categories,
        },
        'percentage_increase': percentage_increase,
    }


    
    return render(request, 'statistik/statistik kerusakan.html',context)

@login_required(login_url=settings.LOGIN_URL)
def statistikpembangunan(request):
    total_entries = Pembangunan.objects.filter(status__tipestatus='PEMBANGUNAN').count()
    max_entries = 100  # Maksimal kerusakan per tahun

    # Hitung persentase pertambahan data
    percentage_increase = (total_entries / max_entries) * 100
    data = Pembangunan.objects.filter(status__tipestatus='PEMBANGUNAN').values('jenis_perlengkapan__jenis_perlengkapan', 'konstruksi_selesai').annotate(data_count=Count('jenis_perlengkapan__jenis_perlengkapan')).order_by('konstruksi_selesai')

    data_line = defaultdict(dict)
    jenis_perlengkapan_choices = Perlengkapan_jalan.objects.values_list('jenis_perlengkapan', flat=True)

    for entry in data:
        jenis_perlengkapan = entry['jenis_perlengkapan__jenis_perlengkapan']
        konstruksi_selesai = entry['konstruksi_selesai'].strftime('%Y-%m-%d')
        data_count = entry['data_count']

        if jenis_perlengkapan not in data_line:
            data_line[jenis_perlengkapan] = {}

        data_line[jenis_perlengkapan][konstruksi_selesai] = data_count

    chart_categories = sorted(set([tanggal for data in data_line.values() for tanggal in data.keys()]))

    chart_data = [{'name': jenis, 'data': [data.get(tanggal, 0) for tanggal in chart_categories]} for jenis, data in data_line.items()]

    data_line_a = Pembangunan.objects.filter(status__tipestatus='PEMBANGUNAN').annotate(date_count=Count('konstruksi_selesai')).values('konstruksi_selesai', 'date_count')

# Membuat dictionary untuk menyimpan data sesuai dengan tanggal
    data_line_plot = {}
    for entry in data_line_a:
        date = entry['konstruksi_selesai'].strftime('%Y-%m-%d')
        data_line_plot[date] = data_line_plot.get(date, 0) + entry['date_count']

    # Mendapatkan rentang tanggal yang diinginkan
    start_date = min(data_line_plot.keys())
    end_date = max(data_line_plot.keys())
    date_range = [start_date]
    current_date = datetime.strptime(start_date, '%Y-%m-%d')
    while current_date < datetime.strptime(end_date, '%Y-%m-%d'):
        current_date += timedelta(days=1)
        date_range.append(current_date.strftime('%Y-%m-%d'))

    # Mengisi data yang kosong dengan nilai nol (0)
    chart_data_line = [{'x': datetime.strptime(date, '%Y-%m-%d').strftime('%d %b %Y'), 'y': data_line_plot.get(date, 0)} for date in date_range]
    
    context = {
        'chart_data_line':chart_data_line,
        'percentage_increase':percentage_increase,
        'chart_data': chart_data,
        'chart_categories': chart_categories,
    }
    return render(request, 'statistik/statistik pembangunan.html',context)

@login_required(login_url=settings.LOGIN_URL)
def statistikpengajuan(request):
    data = Pengajuan.objects.values('jenis_perlengkapan__jenis_perlengkapan', 'tanggal_ajukan').annotate(data_count=Count('jenis_perlengkapan__jenis_perlengkapan'))
    

    data_line = defaultdict(lambda: defaultdict(int))
    for entry in data:
        jenis_perlengkapan = entry['jenis_perlengkapan__jenis_perlengkapan']
        tanggal_ajukan = entry['tanggal_ajukan'].strftime('%Y-%m-%d')
        data_count = entry['data_count']

        data_line[jenis_perlengkapan][tanggal_ajukan] += data_count

    chart_categories = sorted(set([tanggal for data in data_line.values() for tanggal in data.keys()]))

    chart_data = [{'name': jenis_perlengkapan, 'data': [data.get(tanggal, 0) for tanggal in chart_categories]} for jenis_perlengkapan, data in data_line.items()]

    # untuk chart lain 
    total_entries = Pengajuan.objects.count()
    max_entries = 100  # Maksimal kerusakan per tahun

    # Hitung persentase pertambahan data
    percentage_increase = (total_entries / max_entries) * 100
    # data_line = Pengajuan.objects.values('tanggal_ajukan').annotate(total_count=Count('tanggal_ajukan'))
    data_line_a = Pengajuan.objects.annotate(date_count=Count('tanggal_ajukan')).values('tanggal_ajukan', 'date_count').order_by('tanggal_ajukan')
    data_line_plot = {}
    for entry in data_line_a:
        date = entry['tanggal_ajukan'].strftime('%Y-%m-%d')
        data_line_plot[date] = data_line_plot.get(date, 0) + entry['date_count']

    # Mendapatkan rentang tanggal yang diinginkan
    start_date = min(data_line_plot.keys())
    end_date = max(data_line_plot.keys())
    date_range = [start_date]
    current_date = datetime.strptime(start_date, '%Y-%m-%d')
    while current_date < datetime.strptime(end_date, '%Y-%m-%d'):
        current_date += timedelta(days=1)
        date_range.append(current_date.strftime('%Y-%m-%d'))

    # Mengisi data yang kosong dengan nilai nol (0)
    chart_data_line = [{'x': datetime.strptime(date, '%Y-%m-%d').strftime('%d %b %Y'), 'y': data_line_plot.get(date, 0)} for date in date_range]
        
    context = {
        'chart_data_line':chart_data_line,
        'percentage_increase':percentage_increase,
        'chart_data': {
            'series': chart_data,
            'categories': chart_categories,
        },
    }
    return render(request, 'statistik/statistik pengajuan.html',context)

@login_required(login_url=settings.LOGIN_URL)
def statistikperencanaan(request):
    
    total_entries = Pembangunan.objects.filter(status__tipestatus='PERENCANAAN').count()
    max_entries = 100  # Maksimal kerusakan per tahun

    # Hitung persentase pertambahan data
    percentage_increase = (total_entries / max_entries) * 100
    data = Pembangunan.objects.filter(status__tipestatus='PERENCANAAN').values('jenis_perlengkapan__jenis_perlengkapan', 'konstruksi_selesai').annotate(data_count=Count('jenis_perlengkapan__jenis_perlengkapan')).order_by('konstruksi_selesai')

    data_line = defaultdict(dict)
    jenis_perlengkapan_choices = Perlengkapan_jalan.objects.values_list('jenis_perlengkapan', flat=True)

    for entry in data:
        jenis_perlengkapan = entry['jenis_perlengkapan__jenis_perlengkapan']
        konstruksi_selesai = entry['konstruksi_selesai'].strftime('%Y-%m-%d')
        data_count = entry['data_count']

        if jenis_perlengkapan not in data_line:
            data_line[jenis_perlengkapan] = {}

        data_line[jenis_perlengkapan][konstruksi_selesai] = data_count

    chart_categories = sorted(set([tanggal for data in data_line.values() for tanggal in data.keys()]))

    chart_data = [{'name': jenis, 'data': [data.get(tanggal, 0) for tanggal in chart_categories]} for jenis, data in data_line.items()]

    data_line_a = Pembangunan.objects.filter(status__tipestatus='PERENCANAAN').annotate(date_count=Count('konstruksi_selesai')).values('konstruksi_selesai', 'date_count')

# Membuat dictionary untuk menyimpan data sesuai dengan tanggal
    data_line_plot = {}
    for entry in data_line_a:
        date = entry['konstruksi_selesai'].strftime('%Y-%m-%d')
        data_line_plot[date] = data_line_plot.get(date, 0) + entry['date_count']

    # Mendapatkan rentang tanggal yang diinginkan
    start_date = min(data_line_plot.keys())
    end_date = max(data_line_plot.keys())
    date_range = [start_date]
    current_date = datetime.strptime(start_date, '%Y-%m-%d')
    while current_date < datetime.strptime(end_date, '%Y-%m-%d'):
        current_date += timedelta(days=1)
        date_range.append(current_date.strftime('%Y-%m-%d'))

    # Mengisi data yang kosong dengan nilai nol (0)
    chart_data_line = [{'x': datetime.strptime(date, '%Y-%m-%d').strftime('%d %b %Y'), 'y': data_line_plot.get(date, 0)} for date in date_range]
    
    context = {
        'chart_data_line':chart_data_line,
        'percentage_increase':percentage_increase,
        'chart_data': chart_data,
        'chart_categories': chart_categories,
    }
    return render(request, 'statistik/statistik perencanaan.html',context)





def maps(request):
    pembangunan_perencanaan = Pembangunan.objects.filter(status__tipestatus='PEMBANGUNAN').exclude(location=None)
    markers = []
    for pembangunan in pembangunan_perencanaan:
        if pembangunan.location is not None:
            fasilitas = pembangunan.nama_fasilitas
            color = fasilitas.color
            konstruksi_selesai_formatted = pembangunan.konstruksi_selesai.strftime('%b %d, %Y')
            marker = {
                'lat': pembangunan.location.latitude,
                'lng': pembangunan.location.longitude,
                'nama_fasilitas': pembangunan.nama_fasilitas,
                'jenis_perlengkapan': pembangunan.jenis_perlengkapan,
                'konstruksi_selesai': konstruksi_selesai_formatted,
                'ruasjalan': pembangunan.ruasjalan,
                'color': color,
            }
            if pembangunan.gambar:
                marker['gambar'] = pembangunan.gambar.url
            else:
                marker['gambar'] = None
            markers.append(marker)
            
    context = {
        'markers': markers
    }
    return render(request, 'masyarakat/maps.html',context)


def beranda(request):
    return render(request, 'masyarakat/landingpage.html')

# hola
def jenisperlengkapan(request):
    fasilitas_data = Fasilitas_perlengkapan.objects.all()
    perlengkapan_list = Perlengkapan_jalan.objects.all()

    

    context = {
        
        'fasilitas_data': fasilitas_data,
        'perlengkapan_list': perlengkapan_list,
    }
    return render(request, 'masyarakat/jenis perlengkapan.html',context)


# def pengajuan(request):
#     qs = Perlengkapan_jalan.objects.all()
#     return render(request, 'masyarakat/pengajuan.html', {'qs': qs})

def get_json_datafpj(request):
    qs_val = list(Perlengkapan_jalan.objects.values())
    return JsonResponse({'data': qs_val})

def get_jenis_perlengkapan(request):
    jenis_perlengkapan_list = Perlengkapan_jalan.objects.values_list('jenis_perlengkapan', flat=True)
    return JsonResponse(list(jenis_perlengkapan_list), safe=False)

def get_nama_fasilitas(request):
    jenis_perlengkapan = request.GET.get('jenis_perlengkapan')
    nama_fasilitas_list = Fasilitas_perlengkapan.objects.filter(jenis_perlengkapan__jenis_perlengkapan=jenis_perlengkapan).values_list('nama_fasilitas', flat=True)
    return JsonResponse(list(nama_fasilitas_list), safe=False)

def laporkerusakan(request):
    rusak_choices = Rusak.RUSAK_CHOICES
    if request.method == 'POST':
        kerusakan_form = KerusakanForm(request.POST, request.FILES)

        if kerusakan_form.is_valid():
            # Membuat objek Masyarakat dan mengisi kolom-kolomnya
            masyarakat = Masyarakat.objects.create(
                nama=request.POST['nama'],
                notelepon=request.POST['notelepon'],
                alamat=request.POST['alamat']
            )

            # Mengisi kolom-kolom pada objek Pengajuan
            kerusakan = kerusakan_form.save(commit=False)
            kerusakan.masyarakatid = masyarakat
            kerusakan.nama_fasilitas = Fasilitas_perlengkapan.objects.get(nama_fasilitas=request.POST['nama_fasilitas'])
            kerusakan.jenis_perlengkapan = Perlengkapan_jalan.objects.get(jenis_perlengkapan=request.POST['jenis_perlengkapan'])
            kerusakan.deskripsi = request.POST['deskripsi']
            if 'gambar' in request.FILES:
                kerusakan.gambar = request.FILES['gambar']

            # Simpan objek kerusakan
            kerusakan.save()

            # Dapatkan objek Status dengan nilai 'DIAJUKAN'
            tipe_rusak = request.POST.get('tipe_rusak')
            rusak = Rusak.objects.create(tiperusak=tipe_rusak)
            kerusakan.rusak = rusak
            # Menghubungkan objek kerusakan dengan objek status
            

            # Buat objek Location dan hubungkan dengan objek kerusakan
            location = Location.objects.create(
                latitude=request.POST['latitudeA'],
                longitude=request.POST['longitudeB'],
            )
            kerusakan.location = location
            kerusakan.save()

            messages.success(request, 'Data berhasil dilaporkan')
            return redirect('laporkerusakan')  # Ganti 'peta' dengan URL halaman peta

        else:
            errors = kerusakan_form.errors
            print(errors)
            messages.error(request, 'Terjadi kesalahan dalam menambahkan data.')

    else:
        kerusakan_form = KerusakanForm()

    
    
    context={
        'rusak_choices':rusak_choices,
        'kerusakan_form':kerusakan_form,
    }
    return render(request, 'masyarakat/lapor kerusakan.html',context)


def create_pengajuan(request):
    if request.method == 'POST':
        pengajuan_form = PengajuanForm(request.POST)

        if pengajuan_form.is_valid():
            # Membuat objek Masyarakat dan mengisi kolom-kolomnya
            masyarakat = Masyarakat.objects.create(
                nama=request.POST['nama'],
                notelepon=request.POST['notelepon'],
                alamat=request.POST['alamat']
            )

            # Mengisi kolom-kolom pada objek Pengajuan
            pengajuan = pengajuan_form.save(commit=False)
            pengajuan.masyarakatid = masyarakat
            pengajuan.nama_fasilitas = Fasilitas_perlengkapan.objects.get(nama_fasilitas=request.POST['nama_fasilitas'])
            pengajuan.jenis_perlengkapan = Perlengkapan_jalan.objects.get(jenis_perlengkapan=request.POST['jenis_perlengkapan'])
            pengajuan.Fasilitas_khusus = request.POST['Fasilitas_khusus']
            if 'gambar' in request.FILES:
                pengajuan.gambar = request.FILES['gambar']

            # Simpan objek pengajuan
            pengajuan.save()

            # Dapatkan objek Status dengan nilai 'DIAJUKAN'
            status = Status.objects.create(tipestatus='DIAJUKAN')

            # Menghubungkan objek pengajuan dengan objek status
            pengajuan.statuspengajuan = status
            pengajuan.save()

            # Buat objek Location dan hubungkan dengan objek Pengajuan
            location = Location.objects.create(
                latitude=request.POST['latitudeA'],
                longitude=request.POST['longitudeB'],
            )
            pengajuan.location = location
            pengajuan.save()

            messages.success(request, 'Berhasil diajukan')
            return redirect('create_pengajuan')  # Ganti 'peta' dengan URL halaman peta

        else:
            errors = pengajuan_form.errors
            print(errors)
            messages.error(request, 'Terjadi kesalahan dalam menambahkan data.')

    else:
        pengajuan_form = PengajuanForm()

    return render(request, 'masyarakat/pengajuan.html', {'pengajuan_form': pengajuan_form})







# percobaan
@login_required(login_url=settings.LOGIN_URL)
def cadangan(request):
    return render(request, 'masyarakat/cadangan.html')
# percobaan


# khusus akun
@login_required(login_url=settings.LOGIN_URL)
def Account(request):
    admin = Admin.objects.get(username=request.user.username)
    if request.method == 'POST':
        form = EditProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profil berhasil diperbarui.')
            return redirect('Account')
        else:
            errors = form.errors
            print(errors)
            messages.error(request, 'Terjadi kesalahan pada sistem')
    else:
        form = EditProfileForm(instance=admin)

    context = {
        'form': form
    }

    # Menambahkan cek jika 'gambar' adalah file yang valid
    if hasattr(admin, 'gambar') and admin.gambar:
        context['gambar'] = admin.gambar.url
    else:
        context['gambar'] = None

    return render(request, 'core/Akun.html', context)

# khusus akun


# Peta admin

@login_required(login_url=settings.LOGIN_URL)
def petaroutingpembangunan(request,pembangunan_id):
    pembangunan_perencanaan = Pembangunan.objects.filter(status__tipestatus='PEMBANGUNAN',pk=pembangunan_id).exclude(location=None)
    markers = []
    for pembangunan in pembangunan_perencanaan:
        if pembangunan.location is not None:
            fasilitas = pembangunan.nama_fasilitas
            color = fasilitas.color
            konstruksi_selesai_formatted = pembangunan.konstruksi_selesai.strftime('%b %d, %Y')
            marker = {
                'lat': pembangunan.location.latitude,
                'lng': pembangunan.location.longitude,
                'nama_fasilitas': pembangunan.nama_fasilitas,
                'konstruksi_selesai': konstruksi_selesai_formatted,
                'ruasjalan': pembangunan.ruasjalan,
                'color': color,
            }
            if pembangunan.gambar:
                marker['gambar'] = pembangunan.gambar.url
            else:
                marker['gambar'] = None
            markers.append(marker)
            
    context = {
        'markers': markers
    }
    return render(request, 'peta/peta routing pembangunan.html',context)

@login_required(login_url=settings.LOGIN_URL)
def petaroutingperencanaan(request,pembangunan_id):
    pembangunan_perencanaan = Pembangunan.objects.filter(status__tipestatus='PERENCANAAN',pk=pembangunan_id).exclude(location=None)
    markers = []
    for pembangunan in pembangunan_perencanaan:
        if pembangunan.location is not None:
            fasilitas = pembangunan.nama_fasilitas
            color = fasilitas.color
            konstruksi_selesai_formatted = pembangunan.konstruksi_selesai.strftime('%b %d, %Y')
            marker = {
                'lat': pembangunan.location.latitude,
                'lng': pembangunan.location.longitude,
                'nama_fasilitas': pembangunan.nama_fasilitas,
                'konstruksi_selesai': konstruksi_selesai_formatted,
                'ruasjalan': pembangunan.ruasjalan,
                'color': color,
            }
            if pembangunan.gambar:
                marker['gambar'] = pembangunan.gambar.url
            else:
                marker['gambar'] = None
            markers.append(marker)
            
    context = {
        'markers': markers
    }
    return render(request, 'peta/peta routing perencanaan.html',context)

@login_required(login_url=settings.LOGIN_URL)
def petaroutingpengajuan(request,pengajuan_id):
    pengajuan_data = Pengajuan.objects.filter(pk=pengajuan_id, location__isnull=False)
    markers = []

    for pengajuan in pengajuan_data:
        fasilitas = pengajuan.nama_fasilitas
        color = fasilitas.color  # Mengambil warna dari fasilitas perlengkapan

        marker = {
            'lat': pengajuan.location.latitude,
            'lng': pengajuan.location.longitude,
            'nama': pengajuan.masyarakatid.nama,
            'notelepon': pengajuan.masyarakatid.notelepon,
            'alamat': pengajuan.masyarakatid.alamat,
            'nama_fasilitas': pengajuan.nama_fasilitas,
            'color': color,  # Menambahkan warna ke data marker
        }

        if pengajuan.gambar:
            marker['gambar'] = pengajuan.gambar.url
        else:
            marker['gambar'] = None

        markers.append(marker)

    context = {
        'markers': markers
    }

    return render(request, 'peta/peta routing pengajuan.html', context)


@login_required(login_url=settings.LOGIN_URL)
def petaroutingkerusakan(request,kerusakan_id):
    kerusakan_data = Kerusakan.objects.filter(pk=kerusakan_id, location__isnull=False)
    markers = []
    for kerusakan in kerusakan_data:
    
        if kerusakan.location is not None:
            fasilitas = kerusakan.nama_fasilitas
            color = fasilitas.color 
            marker = {
                'lat': kerusakan.location.latitude,
                'lng': kerusakan.location.longitude,
                'nama': kerusakan.masyarakatid.nama,
                'notelepon': kerusakan.masyarakatid.notelepon,
                'alamat': kerusakan.masyarakatid.alamat,
                'nama_fasilitas': kerusakan.nama_fasilitas,
                'tiperusak':kerusakan.rusak.tiperusak,
                'color': color,
            }
            if kerusakan.gambar:
                marker['gambar'] = kerusakan.gambar.url
            else:
                marker['gambar'] = None
            markers.append(marker)
    context = {
        'markers': markers
    }
    return render(request, 'peta/peta routing kerusakan.html', context)

@login_required(login_url=settings.LOGIN_URL)
def petaadkerusakan(request):
    kerusakan_data = Kerusakan.objects.all().exclude(location=None)
    markers = []
    for kerusakan in kerusakan_data:
    
        if kerusakan.location is not None:
            fasilitas = kerusakan.nama_fasilitas
            color = fasilitas.color 
            marker = {
                'lat': kerusakan.location.latitude,
                'lng': kerusakan.location.longitude,
                'nama': kerusakan.masyarakatid.nama,
                'notelepon': kerusakan.masyarakatid.notelepon,
                'alamat': kerusakan.masyarakatid.alamat,
                'nama_fasilitas': kerusakan.nama_fasilitas,
                'tiperusak':kerusakan.rusak.tiperusak,
                'color': color,
            }
            if kerusakan.gambar:
                marker['gambar'] = kerusakan.gambar.url
            else:
                marker['gambar'] = None
            markers.append(marker)
    context = {
        'markers': markers
    }
    return render(request, 'peta/peta kerusakan.html', context)

@login_required(login_url=settings.LOGIN_URL)
def petaadpembangunan(request):
    pembangunan_perencanaan = Pembangunan.objects.filter(status__tipestatus='PEMBANGUNAN').exclude(location=None)
    markers = []
    for pembangunan in pembangunan_perencanaan:
        if pembangunan.location is not None:
            fasilitas = pembangunan.nama_fasilitas
            color = fasilitas.color
            konstruksi_selesai_formatted = pembangunan.konstruksi_selesai.strftime('%b %d, %Y')
            marker = {
                'lat': pembangunan.location.latitude,
                'lng': pembangunan.location.longitude,
                'nama_fasilitas': pembangunan.nama_fasilitas,
                'konstruksi_selesai': konstruksi_selesai_formatted,
                'ruasjalan': pembangunan.ruasjalan,
                'color': color,
            }
            if pembangunan.gambar:
                marker['gambar'] = pembangunan.gambar.url
            else:
                marker['gambar'] = None
            markers.append(marker)
            
    context = {
        'markers': markers
    }
    return render(request, 'peta/peta pembangunan.html',context)

@login_required(login_url=settings.LOGIN_URL)
def petaadpengajuan(request):
    pengajuan_data = Pengajuan.objects.exclude(location=None)
    markers = []

    for pengajuan in pengajuan_data:
        if pengajuan.location is not None:
            fasilitas = pengajuan.nama_fasilitas
            color = fasilitas.color  # Mengambil warna dari fasilitas perlengkapan
            

            marker = {
                'lat': pengajuan.location.latitude,
                'lng': pengajuan.location.longitude,
                'nama': pengajuan.masyarakatid.nama,
                'notelepon': pengajuan.masyarakatid.notelepon,
                'alamat': pengajuan.masyarakatid.alamat,
                'nama_fasilitas': pengajuan.nama_fasilitas,
                'color': color,  # Menambahkan warna ke data marker
            }

            if pengajuan.gambar:
                marker['gambar'] = pengajuan.gambar.url
            else:
                marker['gambar'] = None

            markers.append(marker)

    context = {
        'markers': markers
    }

    return render(request, 'peta/peta pengajuan.html', context)


@login_required(login_url=settings.LOGIN_URL)
def petaadperencanaan(request):
    pembangunan_perencanaan = Pembangunan.objects.filter(status__tipestatus='PERENCANAAN').exclude(location=None)
    markers = []
    for pembangunan in pembangunan_perencanaan:
        if pembangunan.location is not None:
            fasilitas = pembangunan.nama_fasilitas
            color = fasilitas.color
            konstruksi_selesai_formatted = pembangunan.konstruksi_selesai.strftime('%b %d, %Y')
            marker = {
                'lat': pembangunan.location.latitude,
                'lng': pembangunan.location.longitude,
                'nama_fasilitas': pembangunan.nama_fasilitas,
                'konstruksi_selesai': konstruksi_selesai_formatted,
                'ruasjalan': pembangunan.ruasjalan,
                'color': color,
            }
            if pembangunan.gambar:
                marker['gambar'] = pembangunan.gambar.url
            else:
                marker['gambar'] = None
            markers.append(marker)
    context = {
        'markers': markers
    }
    return render(request, 'peta/peta perencanaan.html',context)





# Data tables
@login_required(login_url=settings.LOGIN_URL)
def datafpj(request):
    fasilitas_data = Fasilitas_perlengkapan.objects.all()
    success_messages = messages.get_messages(request)
    success_message = next((m.message for m in success_messages if m.level == messages.SUCCESS), None)
    error_messages = messages.get_messages(request)
    error_message = next((m.message for m in error_messages if m.level == messages.ERROR), None)
    context = {
        'success_message': success_message,
        'error_message': error_message,
        'fasilitas_data': fasilitas_data,
    }
    return render(request, 'datatables/data fpj.html',context)

@login_required(login_url=settings.LOGIN_URL)
def datapembangunan(request):
    pembangunan_data = Pembangunan.objects.filter(status__tipestatus='PEMBANGUNAN')
    success_messages = messages.get_messages(request)
    success_message = next((m.message for m in success_messages if m.level == messages.SUCCESS), None)
    error_messages = messages.get_messages(request)
    error_message = next((m.message for m in error_messages if m.level == messages.ERROR), None)
    context = {
        'pembangunan_data': pembangunan_data,
        'success_message': success_message,
        'error_message': error_message,
    }

    
    return render(request, 'datatables/pembangunan.html',context)

@login_required(login_url=settings.LOGIN_URL)
def datapengajuan(request):
    pengajuan_data = Pengajuan.objects.all()
    success_messages = messages.get_messages(request)
    success_message = next((m.message for m in success_messages if m.level == messages.SUCCESS), None)
    error_messages = messages.get_messages(request)
    error_message = next((m.message for m in error_messages if m.level == messages.ERROR), None)
    context = {
        'pengajuan_data': pengajuan_data,
        'success_message': success_message,
        'error_message': error_message,
    }
    return render(request, 'datatables/pengajuan.html',context)

@login_required(login_url=settings.LOGIN_URL)
def delete_fasilitas(request,fasilitas_id):
    fasilitas = Fasilitas_perlengkapan.objects.get(pk=fasilitas_id)

    if request.method == 'POST':
        try:
            
            # Hapus objek fasilitas
            fasilitas.delete()
            
            messages.success(request, 'Data berhasil dihapus.')
            return redirect('datafpj')
        
        except Exception as e:
            messages.error(request, 'Terjadi kesalahan saat menghapus data.')
            print(str(e))
        context = {
        'fasilitas': fasilitas
    }
    return render(request, 'datatables/data fpj.html',context)

@login_required(login_url=settings.LOGIN_URL)
def delete_pembangunan(request, pembangunan_id):
    pembangunan = Pembangunan.objects.get(pk=pembangunan_id)
    
    # Menampilkan SweetAlert konfirmasi penghapusan
    if request.method == 'POST':
        try:
            # Hapus data yang berelasi
            pembangunan.location.delete()
            pembangunan.status.delete()
            pembangunan.kondisi.delete()
            
            # Hapus objek pembangunan
            pembangunan.delete()
            
            messages.success(request, 'Data berhasil dihapus.')
            return redirect('datapembangunan')
        
        except Exception as e:
            messages.error(request, 'Terjadi kesalahan saat menghapus data.')
            print(str(e))
        context = {
        'pembangunan': pembangunan
    }
    return render(request, 'datatables/pembangunan.html', context)

@login_required(login_url=settings.LOGIN_URL)
def delete_perencanaan(request, pembangunan_id):
    pembangunan = Pembangunan.objects.get(pk=pembangunan_id)
    
    # Menampilkan SweetAlert konfirmasi penghapusan
    if request.method == 'POST':
        try:
            # Hapus data yang berelasi
            pembangunan.location.delete()
            pembangunan.status.delete()
            pembangunan.kondisi.delete()
            
            # Hapus objek pembangunan
            pembangunan.delete()
            
            messages.success(request, 'Data berhasil dihapus.')
            return redirect('dataperencanaan')
        
        except Exception as e:
            messages.error(request, 'Terjadi kesalahan saat menghapus data.')
            print(str(e))
        context = {
        'pembangunan': pembangunan
    }
    return render(request, 'datatables/perencanaan.html', context)


@login_required(login_url=settings.LOGIN_URL)
def delete_kerusakanringan(request, kerusakanid):
    kerusakan = Kerusakan.objects.get(pk=kerusakanid)
    
    if request.method == 'POST':
        try:
            # Hapus data yang berelasi
            kerusakan.location.delete()
            kerusakan.rusak.delete()
            kerusakan.masyarakatid.delete()
            
            # Hapus objek kerusakan
            kerusakan.delete()
            
            messages.success(request, 'Data berhasil dihapus.')
            return redirect('kerusakanringan')
        
        except Exception as e:
            messages.error(request, 'Terjadi kesalahan saat menghapus data.')
            print(str(e))
    
    context = {
        'kerusakan': kerusakan
    }
    return render(request, 'kerusakan/ringan.html', context)

@login_required(login_url=settings.LOGIN_URL)
def delete_kerusakansedang(request, kerusakanid):
    kerusakan = Kerusakan.objects.get(pk=kerusakanid)
    
    if request.method == 'POST':
        try:
            # Hapus data yang berelasi
            kerusakan.location.delete()
            kerusakan.rusak.delete()
            kerusakan.masyarakatid.delete()
            
            # Hapus objek kerusakan
            kerusakan.delete()
            
            messages.success(request, 'Data berhasil dihapus.')
            return redirect('kerusakansedang')
        
        except Exception as e:
            messages.error(request, 'Terjadi kesalahan saat menghapus data.')
            print(str(e))
    
    context = {
        'kerusakan': kerusakan
    }
    return render(request, 'kerusakan/sedang.html', context)

@login_required(login_url=settings.LOGIN_URL)
def delete_kerusakanberat(request, kerusakanid):
    kerusakan = Kerusakan.objects.get(pk=kerusakanid)
    
    if request.method == 'POST':
        try:
            # Hapus data yang berelasi
            kerusakan.location.delete()
            kerusakan.rusak.delete()
            kerusakan.masyarakatid.delete()
            
            # Hapus objek kerusakan
            kerusakan.delete()
            
            messages.success(request, 'Data berhasil dihapus.')
            return redirect('kerusakanberat')
        
        except Exception as e:
            messages.error(request, 'Terjadi kesalahan saat menghapus data.')
            print(str(e))
    
    context = {
        'kerusakan': kerusakan
    }
    return render(request, 'kerusakan/berat.html', context)


@login_required(login_url=settings.LOGIN_URL)
def delete_pengajuan(request, pengajuan_id):
    pengajuan = Pengajuan.objects.get(pk=pengajuan_id)
    
    # Menampilkan SweetAlert konfirmasi penghapusan
    if request.method == 'POST':
        try:
            # Hapus data yang berelasi
            pengajuan.location.delete()
            pengajuan.statuspengajuan.delete()
            pengajuan.masyarakatid.delete()
            
            # Hapus objek Pengajuan
            pengajuan.delete()
            
            messages.success(request, 'Data berhasil dihapus.')
            return redirect('datapengajuan')
        
        except Exception as e:
            messages.error(request, 'Terjadi kesalahan saat menghapus data.')
            print(str(e))
    
    context = {
        'pengajuan': pengajuan
    }
    return render(request, 'datatables/pengajuan.html', context)



@login_required(login_url=settings.LOGIN_URL)
def dataperencanaan(request):
    pembangunan_data = Pembangunan.objects.filter(status__tipestatus='PERENCANAAN')
    success_messages = messages.get_messages(request)
    success_message = next((m.message for m in success_messages if m.level == messages.SUCCESS), None)
    error_messages = messages.get_messages(request)
    error_message = next((m.message for m in error_messages if m.level == messages.ERROR), None)
    context = {
        'pembangunan_data': pembangunan_data,
        'success_message': success_message,
        'error_message': error_message,
    }
    
    return render(request, 'datatables/perencanaan.html',context)

# seleksi
@login_required(login_url=settings.LOGIN_URL)
def seleksidibawah(request):
    pembangunan_data = Pembangunan.objects.filter(status__tipestatus='PEMBANGUNAN',kondisi__tipekondisi='BURUK')
    success_messages = messages.get_messages(request)
    success_message = next((m.message for m in success_messages if m.level == messages.SUCCESS), None)
    error_messages = messages.get_messages(request)
    error_message = next((m.message for m in error_messages if m.level == messages.ERROR), None)
    context = {
        'pembangunan_data': pembangunan_data,
        'success_message': success_message,
        'error_message': error_message,
    }
    return render(request, 'seleksi/dibawah standar.html',context)

@login_required(login_url=settings.LOGIN_URL)
def seleksimemenuhi(request):
    pembangunan_data = Pembangunan.objects.filter(status__tipestatus='PEMBANGUNAN',kondisi__tipekondisi='BAIK')
    success_messages = messages.get_messages(request)
    success_message = next((m.message for m in success_messages if m.level == messages.SUCCESS), None)
    error_messages = messages.get_messages(request)
    error_message = next((m.message for m in error_messages if m.level == messages.ERROR), None)
    context = {
        'pembangunan_data': pembangunan_data,
        'success_message': success_message,
        'error_message': error_message,
    }
    return render(request, 'seleksi/memenuhi standar.html',context)

# pelaporan 
@login_required(login_url=settings.LOGIN_URL)
def pelaporan(request):
    admin = request.user
    context={
        'admin':admin,
    }
    return render(request, 'kerusakan/pelaporan.html',context)

# kerusakan
@login_required(login_url=settings.LOGIN_URL)
def kerusakanberat(request):
    kerusakan_data = Kerusakan.objects.filter(rusak__tiperusak="RUSAK BERAT")
    success_messages = messages.get_messages(request)
    success_message = next((m.message for m in success_messages if m.level == messages.SUCCESS), None)
    error_messages = messages.get_messages(request)
    error_message = next((m.message for m in error_messages if m.level == messages.ERROR), None)
    context = {
        'kerusakan_data':kerusakan_data,
        'success_message': success_message,
        'error_message': error_message,
    }
    return render(request, 'kerusakan/berat.html',context)

@login_required(login_url=settings.LOGIN_URL)
def kerusakansedang(request):
    kerusakan_data = Kerusakan.objects.filter(rusak__tiperusak="RUSAK SEDANG")
    success_messages = messages.get_messages(request)
    success_message = next((m.message for m in success_messages if m.level == messages.SUCCESS), None)
    error_messages = messages.get_messages(request)
    error_message = next((m.message for m in error_messages if m.level == messages.ERROR), None)
    context = {
        'kerusakan_data':kerusakan_data,
        'success_message': success_message,
        'error_message': error_message,
    }
    return render(request, 'kerusakan/sedang.html',context)

@login_required(login_url=settings.LOGIN_URL)
def kerusakanringan(request):
    kerusakan_data = Kerusakan.objects.filter(rusak__tiperusak="RUSAK RINGAN")
    success_messages = messages.get_messages(request)
    success_message = next((m.message for m in success_messages if m.level == messages.SUCCESS), None)
    error_messages = messages.get_messages(request)
    error_message = next((m.message for m in error_messages if m.level == messages.ERROR), None)
    context = {
        'kerusakan_data':kerusakan_data,
        'success_message': success_message,
        'error_message': error_message,
    }
    return render(request, 'kerusakan/ringan.html',context)

# bantuan
@login_required(login_url=settings.LOGIN_URL)
def kontakoperator(request):
    admin = request.user
    context={
        'admin':admin,
    }
    return render(request, 'bantuan/kontak operator.html',context)

@login_required(login_url=settings.LOGIN_URL)
def layanandinas(request):
    admin = request.user
    context={
        'admin':admin,
    }
    return render(request, 'bantuan/layanandinas.html',context)

@login_required(login_url=settings.LOGIN_URL)
def permasalahan(request):
    admin = request.user
    context={
        'admin':admin,
    }
    return render(request, 'bantuan/pengajuan permasalahan.html',context)

# masukkan data buat admin
@login_required(login_url=settings.LOGIN_URL)
def masukdatapembangunan(request):
    kondisi_choices = Kondisi.KONDISI_CHOICES
    status_choices = Status.STATUS_CHOICES

    if request.method == 'POST':
        pembangunan_form = PembangunanForm(request.POST, request.FILES)
        
        if pembangunan_form.is_valid():
            pembangunan = pembangunan_form.save(commit=False)
            pembangunan.nama_fasilitas = Fasilitas_perlengkapan.objects.get(nama_fasilitas=request.POST['nama_fasilitas'])
            pembangunan.jenis_perlengkapan = Perlengkapan_jalan.objects.get(jenis_perlengkapan=request.POST['jenis_perlengkapan'])
            pembangunan.tanggal_bangun = request.POST['tanggal_bangun']
            pembangunan.konstruksi_selesai = request.POST['konstruksi_selesai']
            pembangunan.volume = request.POST['volume']
            pembangunan.ruasjalan = request.POST['RuasJalan']
            pembangunan.deskripsi = request.POST['deskripsi']
            if 'gambar' in request.FILES:
                pembangunan.gambar = request.FILES['gambar']
            
            tipe_kondisi = request.POST['kondisi_id']
            tipe_status = request.POST['status_id']
            
            kondisi = Kondisi.objects.create(tipekondisi=tipe_kondisi)
            status = Status.objects.create(tipestatus=tipe_status)
            
            pembangunan.kondisi = kondisi
            pembangunan.status = status
            
            location = Location.objects.create(
                latitude=request.POST['latitude'],
                longitude=request.POST['longitude'],
            )
            pembangunan.location = location

            pembangunan.save()
            
            messages.success(request, 'Data berhasil ditambahkan')
            return redirect('masukdatapembangunan')  # Ganti 'masukdatapembangunan' dengan URL yang sesuai
            
        else:
            errors = pembangunan_form.errors
            print(errors)
            messages.error(request, 'Terjadi kesalahan dalam menambahkan data.')
    else:
        pembangunan_form = PembangunanForm()
    
    context = {
        'pembangunan_form': pembangunan_form,
        'kondisi_choices': kondisi_choices,
        'status_choices': status_choices,
    }
    
    return render(request, 'create/Pembangunan.html', context)

@login_required(login_url=settings.LOGIN_URL)
def editdatapembangunan(request,pembangunan_id):
    kondisi_choices = Kondisi.KONDISI_CHOICES
    status_choices = Status.STATUS_CHOICES
    pembangunan = Pembangunan.objects.get(pk=pembangunan_id)

    if request.method == 'POST':
        form = PembangunanForm(request.POST, request.FILES,instance=pembangunan)
        if form.is_valid():
            pembangunan = form.save(commit=False)
            pembangunan.nama_fasilitas = Fasilitas_perlengkapan.objects.get(nama_fasilitas=request.POST['nama_fasilitas'])
            pembangunan.jenis_perlengkapan = Perlengkapan_jalan.objects.get(jenis_perlengkapan=request.POST['jenis_perlengkapan'])
            pembangunan.tanggal_bangun = request.POST['tanggal_bangun']
            pembangunan.konstruksi_selesai = request.POST['konstruksi_selesai']
            pembangunan.ruasjalan = request.POST['RuasJalan']
            pembangunan.volume = request.POST['volume']
            pembangunan.deskripsi = request.POST['deskripsi']
            if 'gambar' in request.FILES:
                pembangunan.gambar = request.FILES['gambar']
            
            tipe_kondisi = request.POST['kondisi_id']
            tipe_status = request.POST['status_id']
            
            kondisi = Kondisi.objects.create(tipekondisi=tipe_kondisi)
            status = Status.objects.create(tipestatus=tipe_status)
            
            pembangunan.kondisi = kondisi
            pembangunan.status = status
            
            location = Location.objects.create(
                latitude=request.POST['latitude'],
                longitude=request.POST['longitude'],
            )
            pembangunan.location = location

            pembangunan.save()

            messages.success(request, 'Data berhasil diubah dari data sebelumnya')
            return redirect('datapembangunan')  # Ganti 'nama_halaman' dengan URL halaman yang sesuai
        else:
            messages.error(request, 'Terjadi kesalahan dalam menyimpan data.')
    else:
        form = PembangunanForm(initial={
            'latitude': pembangunan.location.latitude,
            'longitude': pembangunan.location.longitude,
            'kondisi': pembangunan.kondisi.tipekondisi,
            'status': pembangunan.status.tipestatus,
            'deskripsi': pembangunan.deskripsi,
            'volume': pembangunan.volume,
            'tanggal_bangun':pembangunan.tanggal_bangun,
            'konstruksi_selesai': pembangunan.konstruksi_selesai,
            'nama_fasilitas': pembangunan.nama_fasilitas,
            'jenis_perlengkapan': pembangunan.jenis_perlengkapan,
            'ruasjalan':pembangunan.ruasjalan,
        })
        context = {
        'pembangunan': pembangunan,
        'form': form,
        'kondisi_choices': kondisi_choices,
        'status_choices': status_choices,
    }
    
    return render(request, 'create/Edit data.html', context)

    
@login_required(login_url=settings.LOGIN_URL)
def masukdataperencanaan(request):
    kondisi_choices = Kondisi.KONDISI_CHOICES
    status_choices = Status.STATUS_CHOICES

    if request.method == 'POST':
        pembangunan_form = PembangunanForm(request.POST, request.FILES)
        
        if pembangunan_form.is_valid():
            pembangunan = pembangunan_form.save(commit=False)
            pembangunan.nama_fasilitas = Fasilitas_perlengkapan.objects.get(nama_fasilitas=request.POST['nama_fasilitas'])
            pembangunan.jenis_perlengkapan = Perlengkapan_jalan.objects.get(jenis_perlengkapan=request.POST['jenis_perlengkapan'])
            pembangunan.tanggal_bangun = request.POST['tanggal_bangun']
            pembangunan.konstruksi_selesai = request.POST['konstruksi_selesai']
            pembangunan.volume = request.POST['volume']
            pembangunan.ruasjalan = request.POST['RuasJalan']
            pembangunan.deskripsi = request.POST['deskripsi']
            if 'gambar' in request.FILES:
                pembangunan.gambar = request.FILES['gambar']
            
            tipe_kondisi = request.POST['kondisi_id']
            tipe_status = request.POST['status_id']
            
            kondisi = Kondisi.objects.create(tipekondisi=tipe_kondisi)
            status = Status.objects.create(tipestatus=tipe_status)
            
            pembangunan.kondisi = kondisi
            pembangunan.status = status
            
            location = Location.objects.create(
                latitude=request.POST['latitude'],
                longitude=request.POST['longitude'],
            )
            pembangunan.location = location

            pembangunan.save()
            
            messages.success(request, 'Data berhasil ditambahkan')
            return redirect('masukdataperencanaan')  # Ganti 'masukdatapembangunan' dengan URL yang sesuai
            
        else:
            errors = pembangunan_form.errors
            print(errors)
            messages.error(request, 'Terjadi kesalahan dalam menambahkan data.')
    else:
        pembangunan_form = PembangunanForm()
    
    context = {
        'pembangunan_form': pembangunan_form,
        'kondisi_choices': kondisi_choices,
        'status_choices': status_choices,
    }
    return render(request, 'create/Perencanaan.html',context)


@login_required(login_url=settings.LOGIN_URL)
def tambahdataperencanaan(request,pengajuan_id):
    kondisi_choices = Kondisi.KONDISI_CHOICES
    status_choices = Status.STATUS_CHOICES
    pengajuan = Pengajuan.objects.get(pk=pengajuan_id)

    if request.method == 'POST':
        form = PembangunanForm(request.POST, request.FILES)
        if form.is_valid():
            pembangunan = form.save(commit=False)
            pembangunan.nama_fasilitas = Fasilitas_perlengkapan.objects.get(nama_fasilitas=request.POST['nama_fasilitas'])
            pembangunan.jenis_perlengkapan = Perlengkapan_jalan.objects.get(jenis_perlengkapan=request.POST['jenis_perlengkapan'])
            pembangunan.tanggal_bangun = request.POST['tanggal_bangun']
            pembangunan.konstruksi_selesai = request.POST['konstruksi_selesai']
            pembangunan.volume = request.POST['volume']
            pembangunan.ruasjalan = request.POST['RuasJalan']
            pembangunan.deskripsi = request.POST['deskripsi']
            if 'gambar' in request.FILES:
                pembangunan.gambar = request.FILES['gambar']
            
            tipe_kondisi = request.POST['kondisi_id']
            tipe_status = request.POST['status_id']
            
            kondisi = Kondisi.objects.create(tipekondisi=tipe_kondisi)
            status = Status.objects.create(tipestatus=tipe_status)
            
            pembangunan.kondisi = kondisi
            pembangunan.status = status
            
            location = Location.objects.create(
                latitude=request.POST['latitude'],
                longitude=request.POST['longitude'],
            )
            pembangunan.location = location

            pembangunan.save()

            messages.success(request, 'Data berhasil disimpan.')
            return redirect('tambahdataperencanaan')  # Ganti 'nama_halaman' dengan URL halaman yang sesuai
        else:
            messages.error(request, 'Terjadi kesalahan dalam menyimpan data.')
    else:
        form = PembangunanForm(initial={
            'nama': pengajuan.masyarakatid.nama,
            'notelepon':pengajuan.masyarakatid.notelepon,
            'nama_fasilitas': pengajuan.nama_fasilitas.nama_fasilitas,
            'jenis_perlengkapan': pengajuan.jenis_perlengkapan.jenis_perlengkapan,
            'latitude': pengajuan.location.latitude,
            'longitude': pengajuan.location.longitude,
        })

    context = {
        'pengajuan': pengajuan,
        'form': form,
        'kondisi_choices': kondisi_choices,
        'status_choices': status_choices,
    }
    return render(request, 'create/Edit Perencanaan.html', context)



def editdataperlengkapan(request, fasilitas_id):
    fasilitas = get_object_or_404(Fasilitas_perlengkapan, pk=fasilitas_id)
    perlengkapan_list = Perlengkapan_jalan.objects.values_list('jenis_perlengkapan', flat=True)

    if request.method == 'POST':
        fasilitas_form = FasilitasForm(request.POST, request.FILES, instance=fasilitas)
        if fasilitas_form.is_valid():
            nama_fasilitas = fasilitas_form.cleaned_data['nama_fasilitas']
            if Fasilitas_perlengkapan.objects.exclude(pk=fasilitas_id).filter(nama_fasilitas=nama_fasilitas).exists():
                if fasilitas.nama_fasilitas != nama_fasilitas:
                    messages.error(request, 'Nama fasilitas sudah ada di database.')
                else:
                    fasilitas_form.save()
                    messages.success(request, 'Data berhasil diubah')
            else:
                fasilitas_form.save()
                messages.success(request, 'Data berhasil diubah')
                return redirect('datafpj')
        else:
            errors = fasilitas_form.errors
            print(errors)
            messages.error(request, 'Terjadi kesalahan dalam menyimpan data.')
    else:
        # Isi form dengan data yang ada untuk diedit
        fasilitas_form = FasilitasForm(instance=fasilitas)

    context = {
        'fasilitas': fasilitas,
        'fasilitas_form': fasilitas_form,
        'perlengkapan_list': perlengkapan_list,
    }
    return render(request, 'create/Edit data Perlengkapan.html', context)


        
@login_required(login_url=settings.LOGIN_URL)
def masukdataperlengkapan(request):
    perlengkapan_list = Perlengkapan_jalan.objects.values_list('jenis_perlengkapan', flat=True)
    if request.method == 'POST':
        fasilitas_form = FasilitasForm(request.POST, request.FILES)
        if fasilitas_form.is_valid():
            fasilitas = fasilitas_form.save(commit=False)
            fasilitas.nama_fasilitas = request.POST['nama_fasilitas']
            fasilitas.tipekhusus =request.POST['tipekhusus']
            fasilitas.namakhusus =request.POST['namakhusus']
            fasilitas.tanggal_ditambahkan =request.POST['tanggal_bangun']
            fasilitas.volume =request.POST['volume']
            fasilitas.color =request.POST['color']
            fasilitas.jenis_perlengkapan = Perlengkapan_jalan.objects.get(jenis_perlengkapan=request.POST['jenis_perlengkapan'])
            fasilitas.deskripsi =request.POST['deskripsi']
            if 'gambar' in request.FILES:
                fasilitas.gambar = request.FILES['gambar']
            
            fasilitas.save()
            messages.success(request, 'Data berhasil ditambahkan')
            return redirect('masukdataperlengkapan')  # Ganti 'masukdatapembangunan' dengan URL yang sesuai
            
        else:
            errors = fasilitas_form.errors
            print(errors)
            messages.error(request, 'Terjadi kesalahan dalam menambahkan data.')
    else:
        fasilitas_form = FasilitasForm()
            
        
    context = {
        'fasilitas_form': fasilitas_form,
        'perlengkapan_list': perlengkapan_list,
    }
    return render(request, 'create/Perlengkapan jalan.html',context)






# panduan 
@login_required(login_url=settings.LOGIN_URL)
def pemandu(request):
    admin = request.user
    context={
        'admin':admin,
    }
    return render(request, 'panduan/pemandu penambahan.html',context)

@login_required(login_url=settings.LOGIN_URL)
def penggunaan(request):
    admin = request.user
    context={
        'admin':admin,
    }
    return render(request, 'panduan/penggunaan.html',context)

# def login(request):
#     return render(request, 'core/login.html')



# Admin
@login_required(login_url=settings.LOGIN_URL)
def dashboardadmin(request):
    jumlah_data = Pengajuan.objects.count()
    kerusakan = Kerusakan.objects.count()
    data_perencanaan = Pembangunan.objects.filter(status__tipestatus='PERENCANAAN').count()
    data_pembangunan = Pembangunan.objects.filter(status__tipestatus='PEMBANGUNAN').count()
    admin = request.user
    kondisi_choices = {
    'BAIK': 'Kondisi Baik',
    'BURUK': 'Kondisi Buruk',
}
    kondisi_choices = dict(Kondisi.KONDISI_CHOICES)

    data = Pembangunan.objects.values('kondisi__tipekondisi', 'konstruksi_selesai').annotate(data_count=Count('kondisi__tipekondisi'))

    data_linearea = defaultdict(lambda: defaultdict(int))
    for entry in data:
        kondisi_tipe = entry['kondisi__tipekondisi']
        konstruksi_selesai = entry['konstruksi_selesai'].strftime('%Y-%m-%d')
        data_count = entry['data_count']

        data_linearea[kondisi_tipe][konstruksi_selesai] += data_count

    chart_categories = sorted(set([tanggal for data in data_linearea.values() for tanggal in data.keys()]))

    chart_data = [
    {'name': kondisi_choices[kondisi], 'data': [data.get(tanggal, 0) for tanggal in chart_categories]}
    for kondisi, data in data_linearea.items() if kondisi
]
    
    
    # ini bagian notif
    latest_pengajuan = Pengajuan.objects.order_by('-tanggal_ajukan')[:5]

    latest_masyarakat = []
    for pengajuan in latest_pengajuan:
        masyarakat = pengajuan.masyarakatid
        
        latest_masyarakat.append(masyarakat)
        
    latest_kerusakan = Kerusakan.objects.order_by('-tanggal_laporkan')[:5]

    latest_masyarakat_kerusakan = []
    for kerusakandata in latest_kerusakan:
        masyarakat = kerusakandata.masyarakatid
        latest_masyarakat_kerusakan.append(masyarakat)

    pembangunan_perencanaan = Pembangunan.objects.filter(kondisi__tipekondisi='BURUK').exclude(location=None)
    markers = []
    for pembangunan in pembangunan_perencanaan:
        if pembangunan.location is not None:
            fasilitas = pembangunan.nama_fasilitas
            color = fasilitas.color
            konstruksi_selesai_formatted = pembangunan.konstruksi_selesai.strftime('%b %d, %Y')
            marker = {
                'lat': pembangunan.location.latitude,
                'lng': pembangunan.location.longitude,
                'nama_fasilitas': pembangunan.nama_fasilitas,
                'konstruksi_selesai': konstruksi_selesai_formatted,
                'ruasjalan': pembangunan.ruasjalan,
                'kondisi': pembangunan.kondisi.tipekondisi,
                'color': color,
            }
            if pembangunan.gambar:
                marker['gambar'] = pembangunan.gambar.url
            else:
                marker['gambar'] = None
            markers.append(marker)
    
    context = {
        'markers': markers,
        'latest_kerusakan':latest_kerusakan,
        'latest_pengajuan':latest_pengajuan,
        'latest_masyarakat_kerusakan':latest_masyarakat_kerusakan,
        'latest_masyarakat': latest_masyarakat,
        'data_perencanaan' : data_perencanaan,
        'data_pembangunan' : data_pembangunan,
        'jumlah_data': jumlah_data,
        'kerusakan': kerusakan,
        'admin': admin,
        'chart_data': chart_data,
        'chart_categories': chart_categories,
        
        }
    return render(request, 'core/dashboard.html',context)

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('dashboardadmin')
        else:
            error_message = "Invalid username or password"
            return render(request, 'core/login.html', {'error_message': error_message})
    else:
        return render(request, 'core/login.html')
    
def logout_view(request):
    logout(request)
    request.session.flush()
    return redirect('masuk')